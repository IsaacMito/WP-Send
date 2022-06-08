import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class Excel:

    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(path)
        self.sheet: Worksheet = self.wb.active
        self.max_row: int = self.sheet.max_row


    def get_nome(self, r):

        nome: str = str(self.sheet.cell(row=r, column=1).value)

        return nome.split()[0]

    def get_celular(self, r):
        celular: str = str(self.sheet.cell(row=r, column=2).value)

        celular = celular.replace('(', '').replace(')', '').replace(' ', '').replace('+55', '')

        if (len(celular) == 12):
            celular = celular[4:]

        if(len(celular) == 11):
            celular = celular[3:]

        if (len(celular) == 10):
            celular = celular[2:]

        if (len(celular) == 9):
            celular = celular[1:]

        return celular

    def get_status(self, r):
        return self.sheet.cell(row=r, column=3).value

    def salvar(self):
        self.wb.save(self.path)

    def set_status(self, status, r):
        self.sheet.cell(row=r, column=3).value = status
